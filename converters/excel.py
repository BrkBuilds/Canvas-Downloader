import csv
import io
import sys
import time
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelToPDF:
    """Context manager for batch Excel-to-PDF conversion.

    Windows:  Uses COM automation (win32com) with self-healing.
    macOS:    Uses AppleScript via osascript to control Microsoft Excel.

    Design: One COM instance is shared across the batch for speed, but
    self-heals (Quit + re-init) if any individual file crashes the RPC
    channel.  A proactive health check at the start of each convert()
    detects stale COM objects before they cause failures.
    """

    def __init__(self):
        self.app = None
        self._com_pid = None  # PID of the spawned EXCEL.EXE COM process

    # ── lifecycle ──────────────────────────────────────────────────
    def __enter__(self):
        if sys.platform == 'darwin':
            return self  # AppleScript path, no COM needed
        try:
            import pythoncom
            pythoncom.CoInitialize()
        except ImportError:
            pass
        except Exception:
            pass
        self._init_app()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._kill_app()
        # CoUninitialize MUST be called on the same thread as CoInitialize (H-9).
        try:
            import pythoncom
            pythoncom.CoUninitialize()
        except Exception:
            pass

    # ── COM management ─────────────────────────────────────────────
    def _init_app(self):
        """Spin up a fresh, locked-down Excel instance and track its PID."""
        self._com_pid = None
        try:
            import win32com.client
            from engine.office_pid import snapshot_office_pids, find_new_office_pid
            _pre = snapshot_office_pids('EXCEL.EXE')
            self.app = win32com.client.DispatchEx("Excel.Application")
            self.app.Visible = False
            self.app.DisplayAlerts = False
            self.app.EnableEvents = False          # block VBA macros
            try:
                self.app.AutomationSecurity = 3    # msoAutomationSecurityForceDisable
            except Exception:
                pass
            try:
                self.app.Interactive = False        # suppress "Publishing…" dialog
            except Exception:
                pass
            self._com_pid = find_new_office_pid('EXCEL.EXE', _pre)
            logger.debug(f"[COM] Excel started with PID {self._com_pid}")
        except ImportError:
            logger.warning("pywin32 not installed or not on Windows. Excel conversion disabled.")
            self.app = None
        except Exception as e:
            logger.error(f"[COM] Excel init failed: {e}")
            self.app = None

    def _kill_app(self):
        """Forcefully shut down the COM instance (safe to call anytime).

        Quit() alone is NOT enough: when the RPC channel is dead - the exact
        COM Error -2147023174 ("RPC server is unavailable") the self-heal path
        hits - Quit() itself throws and is swallowed, so the EXCEL.EXE we spawned
        is left running as an orphaned, empty Excel window. So after the graceful
        Quit we force-kill the PID we tracked at init, but ONLY if it is still an
        EXCEL.EXE (guards the tiny PID-reuse window after a clean Quit) and ONLY
        that one PID (targeted /PID kill, never a broad /IM) so a workbook the
        user has open in their own Excel is never touched.
        """
        if self.app:
            try:
                self.app.Quit()
            except Exception:
                pass
        self.app = None
        if self._com_pid:
            try:
                from engine.office_pid import kill_office_pid, pid_is_process
                if pid_is_process(self._com_pid, 'EXCEL.EXE'):
                    kill_office_pid(self._com_pid, 'EXCEL.EXE')
            except Exception:
                pass
        self._com_pid = None

    def _is_alive(self) -> bool:
        """Quick COM channel health check - catches stale RPC handles."""
        if not self.app:
            return False
        try:
            _ = self.app.Version  # lightweight roundtrip to Excel
            return True
        except Exception:
            return False

    def _ensure_app(self):
        """Guarantee a live COM instance, reviving if necessary."""
        if not self._is_alive():
            self._kill_app()
            self._init_app()

    # ── AppleScript bridge (macOS) ─────────────────────────────────
    @staticmethod
    def _convert_applescript(src: Path, dst: Path, app_name: str, script: str) -> bool:
        """Delegate to the shared AppleScript bridge (engine/applescript_bridge.py)."""
        from engine.applescript_bridge import run_applescript
        return run_applescript(src, dst, app_name, script)

    def _convert_applescript_excel(self, src: Path, dst: Path) -> bool:
        """Convert an Excel file to PDF via AppleScript on macOS.

        Uses the proven ``open POSIX file`` + ``active workbook`` sequence (this
        is the exact form that has reliably converted Excel→PDF on macOS).

        Robustness rule shared with the PowerPoint/Word converters: the whole
        open→export→close runs inside ``try``; on ANY error we
        ``close active workbook saving no`` and re-raise, so a failed conversion
        can never leave workbooks stacking up open in Excel.

        Prompt handling lives OUTSIDE this script on purpose - putting an
        unverified property like ``set ask to update ... links`` inline is a
        *compile* error (-2741) that ``try`` cannot catch and that takes the
        whole script (and the conversion) down with it. The external-link
        prompt is instead suppressed best-effort, in an isolated osascript, by
        ``prime_office_automation``; macros are handled there via VBAWarnings.
        """
        from engine.applescript_bridge import _as_posix, office_container_stage
        with office_container_stage(src, dst, "Excel") as (s_src, s_dst):
            posix_src = _as_posix(s_src)
            posix_dst = _as_posix(s_dst)
            script = f'''
                tell application "Microsoft Excel"
                    set display alerts to false
                    try
                        open POSIX file "{posix_src}"
                        set theBook to active workbook
                        try
                            tell page setup of active sheet
                                set orientation to landscape
                                set (fit to pages wide) to 1
                                set (fit to pages tall) to false
                            end tell
                        end try
                        save workbook as theBook filename POSIX file "{posix_dst}" file format PDF file format
                        close theBook saving no
                    on error errMsg number errNum
                        try
                            close active workbook saving no
                        end try
                        error errMsg number errNum
                    end try
                end tell
            '''
            return self._convert_applescript(s_src, s_dst, "Excel", script)

    # ── conversion ─────────────────────────────────────────────────
    def convert(self, excel_path: str | Path, dst: str | Path | None = None) -> tuple[str | None, str]:
        """Convert *excel_path* to PDF.  Returns ``(pdf_path, "")`` on
        success or ``(None, error_string)`` on failure."""

        src = Path(excel_path).resolve()
        from pathlib import Path as _P
        # H-7: honour an explicit target (ownership-resolved by the caller).
        dst = _P(dst) if dst is not None else src.with_suffix(".pdf")

        # macOS: AppleScript bridge
        if sys.platform == 'darwin':
            if self._convert_applescript_excel(src, dst):
                src.unlink(missing_ok=True)
                return str(dst), ""
            from engine.applescript_bridge import get_last_error
            _last = get_last_error()
            return None, (_last[1] if _last else "AppleScript conversion failed (unknown error)")

        # Windows: COM automation with path shadowing
        # Proactive health check - catches the "alternating failure" pattern
        # where the PREVIOUS export silently corrupted the COM channel.
        self._ensure_app()
        if not self.app:
            return None, "Excel COM application could not be initialized."

        import threading as _th
        from shared.helpers import office_safe_path
        from engine.office_pid import kill_office_pid

        _COM_TIMEOUT_SECONDS = 180

        with office_safe_path(src, dst=dst) as (safe_src, safe_pdf, true_pdf):
            abs_excel = str(safe_src)
            abs_pdf = str(safe_pdf)
            wb = None
            _timed_out = _th.Event()
            _pid = self._com_pid  # capture before timer fires

            def _on_timeout():
                _timed_out.set()
                logger.error(
                    f"[COM Timeout] Excel hung >{_COM_TIMEOUT_SECONDS}s "
                    f"on {src.name}. Killing PID {_pid or 'unknown'}."
                )
                kill_office_pid(_pid or 0, 'EXCEL.EXE')

            _timer = _th.Timer(_COM_TIMEOUT_SECONDS, _on_timeout)
            _timer.start()
            try:
                wb = self.app.Workbooks.Open(abs_excel, UpdateLinks=0, ReadOnly=True)
                time.sleep(0.3)  # let COM settle

                # Best-effort page-setup: landscape, fit-to-width, zero margins.
                for sheet in wb.Worksheets:
                    try:
                        sheet.PageSetup.Zoom = False
                        sheet.PageSetup.FitToPagesWide = 1
                        sheet.PageSetup.FitToPagesTall = False
                        sheet.PageSetup.Orientation = 2        # xlLandscape
                        sheet.PageSetup.LeftMargin = 0.0
                        sheet.PageSetup.RightMargin = 0.0
                        sheet.PageSetup.TopMargin = 0.0
                        sheet.PageSetup.BottomMargin = 0.0
                    except Exception:
                        pass

                # 0 = xlTypePDF. IncludeDocProperties=True avoids a known Office
                # bug where omitting it triggers a phantom "Save as" dialog when
                # AutomationSecurity is locked down (msoAutomationSecurityForceDisable).
                wb.ExportAsFixedFormat(
                    0, abs_pdf,
                    IncludeDocProperties=True,
                    IgnorePrintAreas=False,
                )
                time.sleep(0.3)

                wb.Close(SaveChanges=False)
                wb = None
                _timer.cancel()
                time.sleep(0.2)

                # ExportAsFixedFormat goes through a printer driver and can
                # return normally having written nothing at all. Deleting the
                # spreadsheet on that basis destroyed the user's only copy, so
                # the PDF has to be shown to exist and be a real PDF first.
                from converters.verify import pdf_looks_real
                _ok, _why = pdf_looks_real(safe_pdf)
                if not _ok:
                    logger.error(
                        f"[COM Converter] Excel reported success for {src.name} "
                        f"but {_why}; keeping the original."
                    )
                    return None, f"Excel reported success but {_why}"

                # Remove the original spreadsheet (from the true long path)
                src.unlink(missing_ok=True)
                # Return the true long-path PDF location (context manager moves it back)
                return str(true_pdf), ""

            except Exception as e:
                _timer.cancel()
                error_msg = str(e)

                if wb is not None:
                    try:
                        wb.Close(SaveChanges=False)
                    except Exception:
                        pass

                # SELF-HEAL: assume the COM channel is dead
                self._kill_app()
                self._init_app()

                if _timed_out.is_set():
                    return None, f"Conversion timed out after {_COM_TIMEOUT_SECONDS}s (Excel stopped responding)"
                return None, f"COM Error: {error_msg}"


class ExcelToData:
    """Context manager for batch Excel-to-structured-text extraction.

    Uses openpyxl (cross-platform, no COM/AppleScript dependency) to produce a
    ``<filename>_Data.txt`` sidecar per workbook with:
    - Coordinate grid (A/B/C columns, 1/2/3 rows) matching the companion PDF
    - Formula annotations: ``250 [Formula: =B2*C2]``
    - Merged cell values repeated across the full merged range
    - Hidden row/column markers: ``[HIDDEN]``
    - In-cell newlines converted to ``<br>`` to preserve CSV line integrity

    The original ``.xlsx`` is intentionally **NOT** deleted.  If the user also
    has Excel→PDF enabled, that converter handles deletion.

    Note: openpyxl reads cached formula results (``data_only=True``).  Values
    reflect what was computed when the file was last saved in Excel.
    """

    _META_CONTEXT = (
        "META-CONTEXT: This document contains structured data extracted from a Microsoft Excel workbook. "
        "Each sheet is separated by a markdown header (### Sheet: [Name]). "
        "Data is formatted as CSV with a coordinate grid: the first row contains column letters (A, B, C...) "
        "and the first column contains row numbers (1, 2, 3...) that match the companion PDF. "
        "Cell annotations: [Formula: =...] shows the underlying formula for calculated cells; "
        "[HIDDEN] marks rows or columns that were hidden in the original file. "
        "Merged cell values are repeated across the full merged range. "
        "Percentage-formatted cells show the display value with a % suffix (e.g. 15.5%). "
        "Date cells are formatted as YYYY-MM-DD. Boolean cells show TRUE or FALSE. "
        "Empty cells are blank."
    )

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    # ── helpers ────────────────────────────────────────────────────

    @staticmethod
    def _build_merged_value_map(ws_val) -> dict:
        """Map (row, col) → (value, number_format) for all non-top-left cells in merged ranges."""
        merged = {}
        try:
            for rng in ws_val.merged_cells.ranges:
                top_cell = ws_val.cell(rng.min_row, rng.min_col)
                top_val = top_cell.value
                top_fmt = top_cell.number_format or 'General'
                for row in range(rng.min_row, rng.max_row + 1):
                    for col in range(rng.min_col, rng.max_col + 1):
                        if row == rng.min_row and col == rng.min_col:
                            continue
                        merged[(row, col)] = (top_val, top_fmt)
        except Exception:
            pass
        return merged

    @staticmethod
    def _format_value(raw_val, number_format: str) -> str:
        """Convert a raw openpyxl cell value to a display string with type awareness."""
        import datetime
        if isinstance(raw_val, bool):
            return 'TRUE' if raw_val else 'FALSE'
        if isinstance(raw_val, datetime.datetime):
            if raw_val.hour == 0 and raw_val.minute == 0 and raw_val.second == 0:
                return raw_val.strftime('%Y-%m-%d')
            return raw_val.strftime('%Y-%m-%d %H:%M')
        if isinstance(raw_val, datetime.date):
            return raw_val.strftime('%Y-%m-%d')
        if isinstance(raw_val, (int, float)) and '%' in (number_format or ''):
            return f"{raw_val * 100:.4g}%"
        return str(raw_val)

    @staticmethod
    def _hidden_rows(ws) -> set:
        try:
            return {i for i, rd in ws.row_dimensions.items() if rd.hidden}
        except Exception:
            return set()

    @staticmethod
    def _hidden_cols(ws) -> set:
        try:
            from openpyxl.utils import column_index_from_string
            return {column_index_from_string(letter) for letter, cd in ws.column_dimensions.items() if cd.hidden}
        except Exception:
            return set()

    @staticmethod
    def _clean(text: str) -> str:
        """Strip CR and replace LF with <br> so in-cell newlines don't break CSV lines."""
        return text.replace('\r', '').replace('\n', ' <br> ')

    def _extract_sheet(self, ws_val, ws_form) -> list[list[str]]:
        """Return one sheet as a 2D list of strings with coordinate grid header."""
        from openpyxl.utils import get_column_letter

        min_r, min_c = ws_val.min_row, ws_val.min_column
        max_r, max_c = ws_val.max_row, ws_val.max_column

        if min_r is None:
            return []

        # Guard against inflated dimension attributes (Excel sometimes writes max_row=1048576
        # for sheets with only formatting applied to full columns, causing catastrophic iteration).
        max_r = min(max_r, 20000)
        max_c = min(max_c, 1000)

        merged = self._build_merged_value_map(ws_val)
        hidden_row_set = self._hidden_rows(ws_val)
        hidden_col_set = self._hidden_cols(ws_val)

        rows_out = []

        # Coordinate header row: ["", "A", "B", "C", ...]
        header = [""] + [get_column_letter(c) for c in range(min_c, max_c + 1)]
        rows_out.append(header)

        for row_idx in range(min_r, max_r + 1):
            row_hidden = row_idx in hidden_row_set
            row_label = f"{row_idx} [HIDDEN]" if row_hidden else str(row_idx)
            row_data = [row_label]

            for col_idx in range(min_c, max_c + 1):
                col_hidden = col_idx in hidden_col_set

                merged_entry = merged.get((row_idx, col_idx))
                if merged_entry is not None:
                    raw_val, cell_fmt = merged_entry
                else:
                    cell = ws_val.cell(row=row_idx, column=col_idx)
                    raw_val = cell.value
                    cell_fmt = cell.number_format or 'General'

                form_val = ws_form.cell(row=row_idx, column=col_idx).value
                has_formula = isinstance(form_val, str) and form_val.startswith('=')

                if raw_val is None and not has_formula:
                    cell_str = "[HIDDEN]" if col_hidden else ""
                elif raw_val is None:
                    # Formula with no cached value
                    cell_str = f"[Formula: {self._clean(form_val)}]"
                    if col_hidden:
                        cell_str += " [HIDDEN]"
                else:
                    clean_val = self._clean(self._format_value(raw_val, cell_fmt))
                    if has_formula:
                        cell_str = f"{clean_val} [Formula: {self._clean(form_val)}]"
                    else:
                        cell_str = clean_val
                    if col_hidden:
                        cell_str += " [HIDDEN]"

                row_data.append(cell_str)

            rows_out.append(row_data)

        return rows_out

    # ── conversion ─────────────────────────────────────────────────

    def convert(self, excel_path: str | Path) -> tuple[str | None, str]:
        """Extract data from *excel_path* into a ``_Data.txt`` sidecar.

        Returns ``(data_txt_path, "")`` on success
        or ``(None, error_string)`` on failure.
        """
        try:
            import openpyxl
        except ImportError:
            return None, "openpyxl is not installed (run: pip install openpyxl)"

        from shared.helpers import office_safe_path

        src = Path(excel_path).resolve()
        dst = src.with_name(src.stem + "_Data.txt")

        # office_safe_path copies the file to a short temp path for paths >240 chars
        # (Win32 COM limitation).  For typical paths it is a zero-cost pass-through.
        # Using it here avoids the \\?\ prefix that make_long_path adds unconditionally,
        # which zipfile/openpyxl can choke on depending on the Windows configuration.
        with office_safe_path(src) as (safe_src, _, _):
            try:
                wb_val = openpyxl.load_workbook(str(safe_src), data_only=True)
                wb_form = openpyxl.load_workbook(str(safe_src), data_only=False)
            except Exception as e:
                return None, f"Failed to open workbook: {e}"

        # Both workbooks are fully in memory; temp copy (if any) is now released.
        from openpyxl.chartsheet import Chartsheet
        sheet_sections = []
        for sheet_name in wb_val.sheetnames:
            try:
                if isinstance(wb_val[sheet_name], Chartsheet):
                    continue
                rows = self._extract_sheet(wb_val[sheet_name], wb_form[sheet_name])
                if rows:
                    sheet_sections.append((sheet_name, rows))
            except Exception as e:
                logger.warning(f"[openpyxl] Skipping sheet '{sheet_name}': {e}")

        wb_val.close()
        wb_form.close()

        if not sheet_sections:
            return None, "No sheets with data found in workbook."

        try:
            with open(str(dst), 'w', encoding='utf-8', newline='') as f:
                f.write(self._META_CONTEXT + "\n\n")
                for sheet_name, rows in sheet_sections:
                    f.write(f"### Sheet: {sheet_name}\n")
                    buf = io.StringIO()
                    writer = csv.writer(buf, lineterminator='\n')
                    for row in rows:
                        writer.writerow(row)
                    f.write(buf.getvalue())
                    f.write("\n\n")
        except PermissionError:
            return None, "Data sidecar in use by another program"
        except Exception as e:
            return None, f"Failed to write data file: {e}"

        return str(dst), ""


