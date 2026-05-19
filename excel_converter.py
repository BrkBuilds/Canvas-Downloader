import csv
import io
import sys
import time
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelToPDF:
    """Context manager for batch Excel-to-PDF conversion.

    Windows:  Uses COM automation (win32com) with self-healing.
    macOS:    Uses AppleScript via osascript to control Microsoft Excel.

    Design: One COM instance is shared across the batch for speed, but
    self-heals (Quit + re-init) if any individual file crashes the RPC
    channel.  A proactive health check at the start of each convert()
    detects stale COM objects before they cause failures.
    """

    def __init__(self):
        self.app = None

    # ── lifecycle ──────────────────────────────────────────────────
    def __enter__(self):
        if sys.platform == 'darwin':
            return self  # AppleScript path, no COM needed
        try:
            import pythoncom
            pythoncom.CoInitialize()
        except ImportError:
            pass
        except Exception:
            pass
        self._init_app()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._kill_app()
        # CoUninitialize MUST be called on the same thread as CoInitialize (H-9).
        try:
            import pythoncom
            pythoncom.CoUninitialize()
        except Exception:
            pass

    # ── COM management ─────────────────────────────────────────────
    def _init_app(self):
        """Spin up a fresh, locked-down Excel instance."""
        try:
            import win32com.client
            self.app = win32com.client.DispatchEx("Excel.Application")
            self.app.Visible = False
            self.app.DisplayAlerts = False
            self.app.EnableEvents = False          # block VBA macros
            try:
                self.app.AutomationSecurity = 3    # msoAutomationSecurityForceDisable
            except Exception:
                pass
            try:
                self.app.Interactive = False        # suppress "Publishing…" dialog
            except Exception:
                pass
        except ImportError:
            logger.warning("pywin32 not installed or not on Windows. Excel conversion disabled.")
            self.app = None
        except Exception as e:
            logger.error(f"[COM] Excel init failed: {e}")
            self.app = None

    def _kill_app(self):
        """Forcefully shut down the COM instance (safe to call anytime)."""
        if self.app:
            try:
                self.app.Quit()
            except Exception:
                pass
        self.app = None

    def _is_alive(self) -> bool:
        """Quick COM channel health check - catches stale RPC handles."""
        if not self.app:
            return False
        try:
            _ = self.app.Version  # lightweight roundtrip to Excel
            return True
        except Exception:
            return False

    def _ensure_app(self):
        """Guarantee a live COM instance, reviving if necessary."""
        if not self._is_alive():
            self._kill_app()
            self._init_app()

    # ── AppleScript bridge (macOS) ─────────────────────────────────
    @staticmethod
    def _convert_applescript(src: Path, dst: Path, app_name: str, script: str) -> bool:
        """Delegate to the shared AppleScript bridge (engine/applescript_bridge.py)."""
        from engine.applescript_bridge import run_applescript
        return run_applescript(src, dst, app_name, script)

    def _convert_applescript_excel(self, src: Path, dst: Path) -> bool:
        """Convert an Excel file to PDF via AppleScript on macOS."""
        from engine.applescript_bridge import _as_posix
        posix_src = _as_posix(src)
        posix_dst = _as_posix(dst)
        script = f'''
            tell application "Microsoft Excel"
                set display alerts to false
                open POSIX file "{posix_src}"
                set theBook to active workbook
                try
                    tell page setup of active sheet
                        set orientation to landscape
                        set (fit to pages wide) to 1
                        set (fit to pages tall) to false
                    end tell
                end try
                save as theBook filename POSIX file "{posix_dst}" file format PDF
                close theBook saving no
            end tell
        '''
        return self._convert_applescript(src, dst, "Excel", script)

    # ── conversion ─────────────────────────────────────────────────
    def convert(self, excel_path: str | Path) -> tuple[str | None, str]:
        """Convert *excel_path* to PDF.  Returns ``(pdf_path, "")`` on
        success or ``(None, error_string)`` on failure."""

        src = Path(excel_path).resolve()
        dst = src.with_suffix(".pdf")

        # macOS: AppleScript bridge
        if sys.platform == 'darwin':
            if self._convert_applescript_excel(src, dst):
                src.unlink(missing_ok=True)
                return str(dst), ""
            return None, "AppleScript conversion failed (is Microsoft Excel installed?)"

        # Windows: COM automation with path shadowing
        # Proactive health check - catches the "alternating failure" pattern
        # where the PREVIOUS export silently corrupted the COM channel.
        self._ensure_app()
        if not self.app:
            return None, "Excel COM application could not be initialized."

        from ui_helpers import office_safe_path

        with office_safe_path(src) as (safe_src, safe_pdf, true_pdf):
            abs_excel = str(safe_src)
            abs_pdf = str(safe_pdf)
            wb = None

            try:
                wb = self.app.Workbooks.Open(abs_excel, UpdateLinks=0, ReadOnly=True)
                time.sleep(0.3)  # let COM settle

                # Best-effort page-setup: landscape, fit-to-width, zero margins.
                for sheet in wb.Worksheets:
                    try:
                        sheet.PageSetup.Zoom = False
                        sheet.PageSetup.FitToPagesWide = 1
                        sheet.PageSetup.FitToPagesTall = False
                        sheet.PageSetup.Orientation = 2        # xlLandscape
                        sheet.PageSetup.LeftMargin = 0.0
                        sheet.PageSetup.RightMargin = 0.0
                        sheet.PageSetup.TopMargin = 0.0
                        sheet.PageSetup.BottomMargin = 0.0
                    except Exception:
                        pass

                # 0 = xlTypePDF
                wb.ExportAsFixedFormat(0, abs_pdf)
                time.sleep(0.3)

                wb.Close(SaveChanges=False)
                wb = None
                time.sleep(0.2)

                # Remove the original spreadsheet (from the true long path)
                src.unlink(missing_ok=True)
                # Return the true long-path PDF location (context manager moves it back)
                return str(true_pdf), ""

            except Exception as e:
                error_msg = str(e)

                if wb is not None:
                    try:
                        wb.Close(SaveChanges=False)
                    except Exception:
                        pass

                # SELF-HEAL: assume the COM channel is dead
                self._kill_app()
                self._init_app()

                return None, f"COM Error: {error_msg}"


class ExcelToData:
    """Context manager for batch Excel-to-structured-text extraction.

    Uses openpyxl (cross-platform, no COM/AppleScript dependency) to produce a
    ``<filename>_Data.txt`` sidecar per workbook with:
    - Coordinate grid (A/B/C columns, 1/2/3 rows) matching the companion PDF
    - Formula annotations: ``250 [Formula: =B2*C2]``
    - Merged cell values repeated across the full merged range
    - Hidden row/column markers: ``[HIDDEN]``
    - In-cell newlines converted to ``<br>`` to preserve CSV line integrity

    The original ``.xlsx`` is intentionally **NOT** deleted.  If the user also
    has Excel→PDF enabled, that converter handles deletion.

    Note: openpyxl reads cached formula results (``data_only=True``).  Values
    reflect what was computed when the file was last saved in Excel.
    """

    _META_CONTEXT = (
        "META-CONTEXT: This document contains structured data extracted from a Microsoft Excel workbook. "
        "Each sheet is separated by a markdown header (### Sheet: [Name]). "
        "Data is formatted as CSV with a coordinate grid: the first row contains column letters (A, B, C...) "
        "and the first column contains row numbers (1, 2, 3...) that match the companion PDF. "
        "Cell annotations: [Formula: =...] shows the underlying formula for calculated cells; "
        "[HIDDEN] marks rows or columns that were hidden in the original file. "
        "Merged cell values are repeated across the full merged range. "
        "Percentage-formatted cells show the display value with a % suffix (e.g. 15.5%). "
        "Date cells are formatted as YYYY-MM-DD. Boolean cells show TRUE or FALSE. "
        "Empty cells are blank."
    )

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    # ── helpers ────────────────────────────────────────────────────

    @staticmethod
    def _build_merged_value_map(ws_val) -> dict:
        """Map (row, col) → (value, number_format) for all non-top-left cells in merged ranges."""
        merged = {}
        try:
            for rng in ws_val.merged_cells.ranges:
                top_cell = ws_val.cell(rng.min_row, rng.min_col)
                top_val = top_cell.value
                top_fmt = top_cell.number_format or 'General'
                for row in range(rng.min_row, rng.max_row + 1):
                    for col in range(rng.min_col, rng.max_col + 1):
                        if row == rng.min_row and col == rng.min_col:
                            continue
                        merged[(row, col)] = (top_val, top_fmt)
        except Exception:
            pass
        return merged

    @staticmethod
    def _format_value(raw_val, number_format: str) -> str:
        """Convert a raw openpyxl cell value to a display string with type awareness."""
        import datetime
        if isinstance(raw_val, bool):
            return 'TRUE' if raw_val else 'FALSE'
        if isinstance(raw_val, datetime.datetime):
            if raw_val.hour == 0 and raw_val.minute == 0 and raw_val.second == 0:
                return raw_val.strftime('%Y-%m-%d')
            return raw_val.strftime('%Y-%m-%d %H:%M')
        if isinstance(raw_val, datetime.date):
            return raw_val.strftime('%Y-%m-%d')
        if isinstance(raw_val, (int, float)) and '%' in (number_format or ''):
            return f"{raw_val * 100:.4g}%"
        return str(raw_val)

    @staticmethod
    def _hidden_rows(ws) -> set:
        try:
            return {i for i, rd in ws.row_dimensions.items() if rd.hidden}
        except Exception:
            return set()

    @staticmethod
    def _hidden_cols(ws) -> set:
        try:
            from openpyxl.utils import column_index_from_string
            return {column_index_from_string(letter) for letter, cd in ws.column_dimensions.items() if cd.hidden}
        except Exception:
            return set()

    @staticmethod
    def _clean(text: str) -> str:
        """Strip CR and replace LF with <br> so in-cell newlines don't break CSV lines."""
        return text.replace('\r', '').replace('\n', ' <br> ')

    def _extract_sheet(self, ws_val, ws_form) -> list[list[str]]:
        """Return one sheet as a 2D list of strings with coordinate grid header."""
        from openpyxl.utils import get_column_letter

        min_r, min_c = ws_val.min_row, ws_val.min_column
        max_r, max_c = ws_val.max_row, ws_val.max_column

        if min_r is None:
            return []

        # Guard against inflated dimension attributes (Excel sometimes writes max_row=1048576
        # for sheets with only formatting applied to full columns, causing catastrophic iteration).
        max_r = min(max_r, 20000)
        max_c = min(max_c, 1000)

        merged = self._build_merged_value_map(ws_val)
        hidden_row_set = self._hidden_rows(ws_val)
        hidden_col_set = self._hidden_cols(ws_val)

        rows_out = []

        # Coordinate header row: ["", "A", "B", "C", ...]
        header = [""] + [get_column_letter(c) for c in range(min_c, max_c + 1)]
        rows_out.append(header)

        for row_idx in range(min_r, max_r + 1):
            row_hidden = row_idx in hidden_row_set
            row_label = f"{row_idx} [HIDDEN]" if row_hidden else str(row_idx)
            row_data = [row_label]

            for col_idx in range(min_c, max_c + 1):
                col_hidden = col_idx in hidden_col_set

                merged_entry = merged.get((row_idx, col_idx))
                if merged_entry is not None:
                    raw_val, cell_fmt = merged_entry
                else:
                    cell = ws_val.cell(row=row_idx, column=col_idx)
                    raw_val = cell.value
                    cell_fmt = cell.number_format or 'General'

                form_val = ws_form.cell(row=row_idx, column=col_idx).value
                has_formula = isinstance(form_val, str) and form_val.startswith('=')

                if raw_val is None and not has_formula:
                    cell_str = "[HIDDEN]" if col_hidden else ""
                elif raw_val is None:
                    # Formula with no cached value
                    cell_str = f"[Formula: {self._clean(form_val)}]"
                    if col_hidden:
                        cell_str += " [HIDDEN]"
                else:
                    clean_val = self._clean(self._format_value(raw_val, cell_fmt))
                    if has_formula:
                        cell_str = f"{clean_val} [Formula: {self._clean(form_val)}]"
                    else:
                        cell_str = clean_val
                    if col_hidden:
                        cell_str += " [HIDDEN]"

                row_data.append(cell_str)

            rows_out.append(row_data)

        return rows_out

    # ── conversion ─────────────────────────────────────────────────

    def convert(self, excel_path: str | Path) -> tuple[str | None, str]:
        """Extract data from *excel_path* into a ``_Data.txt`` sidecar.

        Returns ``(data_txt_path, "")`` on success
        or ``(None, error_string)`` on failure.
        """
        try:
            import openpyxl
        except ImportError:
            return None, "openpyxl is not installed (run: pip install openpyxl)"

        from ui_helpers import office_safe_path

        src = Path(excel_path).resolve()
        dst = src.with_name(src.stem + "_Data.txt")

        # office_safe_path copies the file to a short temp path for paths >240 chars
        # (Win32 COM limitation).  For typical paths it is a zero-cost pass-through.
        # Using it here avoids the \\?\ prefix that make_long_path adds unconditionally,
        # which zipfile/openpyxl can choke on depending on the Windows configuration.
        with office_safe_path(src) as (safe_src, _, _):
            try:
                wb_val = openpyxl.load_workbook(str(safe_src), data_only=True)
                wb_form = openpyxl.load_workbook(str(safe_src), data_only=False)
            except Exception as e:
                return None, f"Failed to open workbook: {e}"

        # Both workbooks are fully in memory; temp copy (if any) is now released.
        sheet_sections = []
        for sheet_name in wb_val.sheetnames:
            try:
                rows = self._extract_sheet(wb_val[sheet_name], wb_form[sheet_name])
                if rows:
                    sheet_sections.append((sheet_name, rows))
            except Exception as e:
                logger.warning(f"[openpyxl] Skipping sheet '{sheet_name}': {e}")

        wb_val.close()
        wb_form.close()

        if not sheet_sections:
            return None, "No sheets with data found in workbook."

        try:
            with open(str(dst), 'w', encoding='utf-8-sig', newline='') as f:
                f.write(self._META_CONTEXT + "\n\n")
                for sheet_name, rows in sheet_sections:
                    f.write(f"### Sheet: {sheet_name}\n")
                    buf = io.StringIO()
                    writer = csv.writer(buf, lineterminator='\n')
                    for row in rows:
                        writer.writerow(row)
                    f.write(buf.getvalue())
                    f.write("\n\n")
        except PermissionError:
            return None, "Data sidecar in use by another program"
        except Exception as e:
            return None, f"Failed to write data file: {e}"

        return str(dst), ""


